import copy
import json
import math
import os
from copy import deepcopy
from pathlib import Path
from typing import Iterable, Union, Generator

import numpy as np
import openpyxl
from oxDNA_analysis_tools.UTILS.RyeReader import Configuration
from Bio.SVDSuperimposer import SVDSuperimposer
from scipy.optimize import linear_sum_assignment

from .dna_particle import DNAParticle, PatchyOriRelation
from .pl.plpatchylib import mgl_to_pl
from .pl.plparticle import PLPatchyParticle
from .pl.plpatch import PLPatch
from .pl.plscene import PLPSimulation
from ipy_oxdna.dna_structure import construct_strands, BASE_BASE, rc, POS_BASE, DNAStructure, DNAStructureStrand

from random import choice
import itertools

from ..patchy.mgl import MGLPatch, MGLScene
from ..patchy_base_particle import PatchyBaseParticle
from ..polycubeutil.polycubesRule import diridx, FACE_NAMES
from ..util import dist, normalize, get_output_dir, selectColor


# todo: MORE PARAMETERS
def patch_positions_to_bind(patch_positions1: Iterable[int],
                            patch_positions2: Iterable[int],
                            conf1: Configuration,
                            conf2: Configuration) -> list[tuple[int, int]]:
    # get patch positions on particle 2 as an array
    pps2 = np.array([conf2.positions[pid] for pid in patch_positions2])

    # get all possible permutations of patch positions on particle 1
    perms = np.array(list(itertools.permutations([conf1.positions[pid] for pid in patch_positions1])))

    # compute vector difference between the particle 2 patch positions and the permutation matrix
    diffs = pps2[np.newaxis, :, :] - perms  # vector difference
    distancesqrd = np.sum(np.multiply(diffs, diffs), axis=(2,))  # dot product, gives us distance squared
    # sum of squares
    sums = np.sum(distancesqrd, axis=(1,))

    bestmatch = np.argmin(sums)
    best_patch1_positions: list[int] = list(itertools.permutations(patch_positions1))[bestmatch]
    return zip(
        best_patch1_positions,
        patch_positions2)

    # pairs = [] # we store pairs of corresponding indicesreverse
    # pp_pos2 = deepcopy(patch_positions2) #to be expendable
    # # go over every patch in the list for 1st particle
    # for patch_id1 in patch_positions1:
    #     # calculate the distances to all patches of particle 2
    #     dists = [dist(conf1.positions[patch_id1] - conf2.positions[id]) for id in pp_pos2]
    #     # figure out the index of the min particle 75
    #     min_dist = min(dists)
    #     id_min = dists.index(min_dist)
    #     # eliminate that index from the positions to scan against
    #     pairs.append(
    #         (patch_id1, pp_pos2.pop(id_min))
    #     )
    # return pairs # the closest patches on the 2 particles provided


def pairwise(iterable):
    # because i have not updated this yet
    a, b = itertools.tee(iterable)
    next(b, None)
    return zip(a, b)


def patch_idxs_to_bind(patch_id_1: int,
                       patch_id_2: int,
                       dna1: DNAParticle,
                       dna2: DNAParticle) -> list[tuple[int, int]]:
    """
    Parameters:
        patch_id_1 (int): indexer for patch 1
        patch_id_2 (int): indexer for patch 2
        dna1 (DNAParticle): first dna particle to bind
        dna2 (DNAParticle): second dna particle to bind

    Return:
        a tuple of pairs of indexes in patch 1 and patch 2 where the indexes are of the strand indexes (i'm so sorry)

    """

    patch_positions1 = dna1.dna_patch_positions(patch_id_1)
    patch_positions2 = dna2.dna_patch_positions(patch_id_2)
    # get patch positions on particle 2 as an array

    # get all possible permutations of patch positions on particle 1
    perms = np.array(list(itertools.permutations([
        patch_positions1[i, :] for i in range(patch_positions1.shape[0])]
    )))

    # compute vector difference between the particle 2 patch positions and the permutation matrix
    diffs = patch_positions2[np.newaxis, :, :] - perms  # vector difference
    distancesqrd = np.sum(np.multiply(diffs, diffs), axis=(2,))  # dot product, gives us distance squared
    # sum of squares
    sums = np.sum(distancesqrd, axis=(1,))

    bestmatch = np.argmin(sums)
    best_patch1_idxs: list[int] = list(itertools.permutations(range(len(patch_positions1))))[bestmatch]
    return zip(
        best_patch1_idxs,
        range(patch_positions2.shape[0]))


class PatchyOrigamiConverter:
    """
    This class facilitates the conversion of a patchy particle model (in MGL format) to
    a set of connected DNA origamis
    """
    color_sequences: dict[Union[int, str], str]
    bondcount: int
    # map where the keys are patchy particle tpye IDs and the values are DNA type particles
    particle_type_map: dict[int, DNAParticle]
    patchy_scene: PLPSimulation
    # mapping where keys are PL particle UIDs and values are DNA particles
    dna_particles: dict[int, DNAParticle]
    padding: float  # manually-entered extra spacing
    dist_ratio: float
    sticky_length: Union[int, None]
    check_strict_rc: bool = True
    scale_factor: Union[float, None] = None
    # "book" of sticky ends which have already been added to the scene, used
    # to find un-added stickies and add them later
    sticky_book: Union[set[tuple[int, int]], None] = set()

    def __init__(self,
                 scene: PLPSimulation,
                 spacer_length: int = 16,
                 particle_delta: float = 1.2,
                 bond_length: float = 0.4,
                 cos_theta_max: float = 0.95,
                 sticky_length: int = 0,
                 padding: float = 1.0,
                 flexable_patch_distances: bool = False,
                 expected_num_edges: int = -1,
                 rel_rms_tolerance=0.15
                 ):
        """
        Initializes this converter using a few required params and a lot of optional fine-tuning params
        Parameters:
            scene: an MGLScene object representing the structure we're making with the particles
            can be a single origami or a dict relating mgl colors to origamis
            spacer_length: length of poly-T spacer between particle and sticky end
            particle_delta: distance beyond which we can assume particles do not interact
            bond_length: distance beyond which we can assume patches do not interact. TODO: replace with some sort of gaussian?
            cos_theta_max: maximim cos(theta) angle between patches (TODO: interact with patch width?)
            sticky_length: length of sticky end sequences
            padding: magic number to get a little bit of space
        """

        # set parameters
        self.spacer_length = spacer_length
        assert spacer_length is not None, "Cannot skip spacer length! To have no spacers, set spacer_length=0"
        self.particle_delta = particle_delta
        self.bond_length = bond_length
        self.cos_theta_max = cos_theta_max
        self.sticky_length = sticky_length

        # prep variables to use when building our structure
        self.color_sequences = {}
        self.color_match_overrides = {}
        self.bondcount = 0
        self.padding = padding
        self.dist_ratio = None
        self.dna_particles = None

        # inputs
        if isinstance(scene, MGLScene):
            scene = mgl_to_pl(scene)
        self.patchy_scene = scene

        self.particle_type_map = dict()

        # optional parameters
        self.flexable_patch_distances = flexable_patch_distances
        self.expected_num_edges = expected_num_edges
        assert not flexable_patch_distances or expected_num_edges > 0
        self.rel_rms_tolerance = rel_rms_tolerance

    def set_track_conf_stickies(self, bNewVal):
        if bNewVal:
            self.sticky_book = set()
        else:
            self.sticky_book = None
    def is_track_conf_stickies(self) ->bool:
        return self.sticky_book is not None

    def get_dna_origami(self,
                        particle_type: Union[str, int, PatchyBaseParticle]) -> DNAParticle:
        if isinstance(particle_type, PatchyBaseParticle):
            return self.get_dna_origami(particle_type.get_type())
        else:
            assert particle_type in self.particle_type_map, f"Particle type {particle_type} not in type map!"
            return self.particle_type_map[particle_type]

    def get_full_conf(self, clusters=False) -> DNAStructure:
        """
        Combines all structures together into a single DNAStructure
        """
        print("merging the topologies")
        particles = self.get_particles()
        if clusters:
            for p in particles:
                p.clear_clusters()
                for b in p.iter_bases():
                    p.assign_base_to_cluster(b.uid, 0)
        merged_conf = sum(particles[1:], start=particles[0])
        return merged_conf

    def get_sticky_length(self) -> float:
        """
        This method is *not* a simple accessor for self.sticky_length!
        If sticky end length has been hard coded it will simply return self.sticky_length
        If sticky length is not hardcoded it will return the average sticky length
        this is important for computing spacing when positioning particles
        """
        if self.sticky_length is not None:
            return self.sticky_length
        else:
            assert len(
                self.color_sequences) > 0, "Can't dynamically calculate sticky end length without assigned sticky ends!"
            return sum([len(seq) for seq in self.color_sequences.values()]) / len(self.color_sequences)

    def check_rms(self, testval: float, dna: DNAParticle) -> float:
        """

        """
        rel_rms = testval / dna.patches_sphere_circumfrance()
        return self.rel_rms_tolerance > rel_rms

    def get_scale_factor(self) -> float:
        """
        Computes the scale factor for the converter, if it hasn't been set manually
        """
        if self.scale_factor is not None:
            return self.scale_factor
        else:
            scale_factor = self.get_dist_ratio() * self.padding
            return scale_factor

    def get_dist_ratio(self) -> float:
        """
        gets the scaling value used to convert the particle scene to the oxDNA conf
        calculates value if it has not already been computed
        Returns:
            a conversion factor in units of oxDNA Units / Scene Units
        """

        if self.dist_ratio is None:
            dist_ratios = []
            dna_distance = (2 * self.spacer_length + self.get_sticky_length()) * BASE_BASE

            # WARNING! this algorithm could become very intense very quickly for large scenes
            for p1, p2 in self.patchy_scene.iter_bound_particles():
                # compute distance between two particles in the scene (mgl or pl)
                particles_distance = dist(p1.position(), p2.position())
                # technically radius() isn't a member of BaseParticle
                # radii: float = p1.radius() + p2.radius()

                p1_dna = self.particle_type_map[p1.get_type()]
                p2_dna = self.particle_type_map[p2.get_type()]

                # compute DNA particle "radius" as determined by average distance between the patch and cms
                p1_dna_rad = p1_dna.center2patch_conf()
                p2_dna_rad = p2_dna.center2patch_conf()

                # distance between cmss of two dna particles
                # should be very close to constant across the system
                dna_distance_2p = p1_dna_rad + p2_dna_rad + dna_distance
                dist_ratio = dna_distance_2p / particles_distance
                dist_ratios.append(dist_ratio)
            dist_ratio = np.mean(dist_ratios)
            # if the distance ratio is NaN, that means there are no bound particles
            # we may however want to continue anyway
            # todo: make this an exception which we can catch
            if np.isnan(dist_ratio):
                print("Warning: scene has no bound particles!")
            elif np.std(dist_ratios) > 0.05 * dist_ratio:
                print(f"Warning: The standard deviation of individual particle distance ratios {np.std(dist_ratios)}"
                      f" is more than 5% of the value of the mean distance ratio {dist_ratio}.")
            self.dist_ratio = dist_ratio
        return self.dist_ratio

    # TODO: write better?
    def color_sequence(self, colorstr: str) -> str:
        # if this color isn't in our color bank
        if colorstr not in self.color_sequences:
            # ...but its matching color is
            if self.get_color_match(colorstr) in self.color_sequences:
                # use the reverse compliment of the matching color sequenece
                self.color_sequences[colorstr] = rc(self.color_sequence(self.get_color_match(colorstr)))
                print(f"Assigning color {colorstr} sequence {self.color_sequences[colorstr]}"
                      f" (reverse compliment of {self.get_color_match(colorstr)})")
            else:
                # if neither our color nor its match is in the color seqs, generate a new sequence
                # todo: smarter?
                assert self.sticky_length is not None, f"Auto-generation of sequences is off but you didn't provide a " \
                                                       f"color sequence for {colorstr}"
                self.color_sequences[colorstr] = "".join(
                    choice(["A", "T", "C", "G"]) for _ in range(self.sticky_length))
                print(f"Assigning color {colorstr} random sequence {self.color_sequences[colorstr]}")

        return self.color_sequences[colorstr]

    def assign_color_sequence(self, color: int, seq: str, update_rc=True):
        """
        Assigns the color given by colorstr the specific sequence sequence specified
        Automatically assigns the corresponding color the reverse compliment
        """
        assert self.sticky_length is None or len(seq) == self.sticky_length, "Incompatible sequence length"
        self.color_sequences[color] = seq
        if update_rc:
            self.color_sequences[-color] = rc(seq)

    def assign_particles(self,
                         dna: DNAParticle,
                         *args: Union[str, PLPatchyParticle, int]):
        """
        Assigns a dna particle to one or more patchy particle types
        """

        # if no particle was provided, assume we're assigning the same DNA particle to
        # all patchy particle types.
        if not len(args):
            patchy_types: list[PLPatchyParticle] = self.patchy_scene.particle_types().particles()
        else:
            patchy_types: list[PLPatchyParticle] = []
            for a in args:
                if isinstance(a, PLPatchyParticle):
                    patchy_types.append(a)
                else:
                    ptype = self.patchy_scene.particle_types().particle(a)
                    patchy_types.append(ptype)
        for patchy_type in patchy_types:
            dna_cpy = deepcopy(dna)
            self.particle_type_map[patchy_type.get_type()] = dna_cpy
            self.link_patchy_particle(patchy_type, dna_cpy)

    def get_color_match(self, color: Union[int, str]) -> Union[int, str]:
        if color in self.color_match_overrides:
            return self.color_match_overrides[color]
        else:
            if isinstance(color, str):
                if color.startswith("dark"):
                    return color[4:]
                else:
                    return f"dark{color}"
            else:
                if not isinstance(color, int):
                    raise TypeError(f"Invalid color type {type(color)}")
                else:
                    return -color

    def get_expected_pad_distance(self) -> float:
        return (self.get_sticky_length() + 2.0 * self.spacer_length) * POS_BASE

    def set_color_match(self,
                        colorstr: str,
                        match: str,
                        reverse_match: bool = True):
        self.color_match_overrides[colorstr] = match
        if reverse_match:
            self.color_match_overrides[match] = colorstr

    def match_patches_to_strands(self,
                                 p: PLPatchyParticle,
                                 dna: DNAParticle) -> PatchyOriRelation:
        """
        Computes a rotation of this DNA particle that makes the patches on that particle line up with the
        3' ends of patch strands.
        Parameters:
            p (PLPatchyParticle): a patchy particle (type) to link
            dna (DNAParticle):
        Returns:
            a dataclass instance describing the relationship between the patchy particle
            and the dna origami
        """
        assert (p.rotmatrix() == np.identity(3)).all(), "Cannot perform match on rotated patchy particle!"

        if self.patchy_scene.particle_types().has_udt_src():
            n_patches_udt = self.patchy_scene.particle_types().get_src().particle(p.type_id()).num_patches()
        else:
            # this is not ideal but we can derive the number of "source patches". badly.
            assert all([len(strand_ids) == len(dna.patch_strand_ids[0]) for strand_ids in dna.patch_strand_ids[1:]])
            assert p.num_patches() % len(dna.patch_strand_ids[0]) == 0
            n_patches_udt = int(p.num_patches() / len(dna.patch_strand_ids[0]))
            raise Exception("Currently do not support converting patchy sets without a unidentate source!")
        # if num patches is less than 3, align teeth individually to prevent gimbal lock problems
        if n_patches_udt < 3:
            best_rms = np.inf
            svd = SVDSuperimposer()

            strand_map = None
            best_rot = None

            for strand_groups in itertools.permutations(dna.patch_strand_ids, n_patches_udt):
                # skip patch normals for this calculation
                # implicit delimination between patch groups
                patches_coords = [patch.position() for patch in p.patches()]

                # construct m2 matrix for pl coords
                m2 = np.stack([*patches_coords, np.zeros(3)]) / dna.scale_factor(p)

                for strand_order in itertools.product(
                        *[itertools.permutations(strand_group, len(strand_group)) for strand_group in strand_groups]):
                    strands = list(itertools.chain.from_iterable(strand_order))
                    strand_coords = [dna.strand_3p(strand_idx).pos for strand_idx in strands]
                    m1 = np.stack([*strand_coords, np.zeros(3)])
                    svd.set(m2, m1)
                    svd.run()
                    rms = svd.get_rms()
                    if rms < best_rms:
                        strand_map = {
                            patch.get_id(): strand for patch, strand in zip(p.patches(), strands)
                        }
                        best_rot, _ = svd.get_rotran()
                        best_rms = rms
            dna.assign_patches_strands(strand_map)
            dna.transform(rot=best_rot)

            relation = PatchyOriRelation(p, dna, best_rot, strand_map, best_rms)
        else:
            # compute best ordering of patches (unidentate formulation)
            relation = self.align_patches(p, dna)
            assert self.check_rms(relation.rms, dna)
            # apply rotation (important for future superimposer fun times
            dna.transform(rot=relation.rot.T)

            # udt_id = id of unidentate expression of patch
            # strand_group_idx = idx in dna of list of DNA strands that correspond to patch
            # we can't try to find the totally optimal strand ordering because for 6 patches of 4 strands each
            # # that's (4!)^6 = too much
            for udt_id, strand_group_idx in relation.perm.items():
                # get strand IDs
                strand_group = dna.patch_strand_ids[strand_group_idx]
                # get the multidentate representation of the patch
                patch_group = self.patchy_scene.particle_types().mdt_rep(udt_id)
                assert all([(abs(p.patch_by_id(patch.get_id()).position() - patch.position()) < 1e-6).all() for patch in
                            patch_group])
                # compute strand mapping
                strand_map: dict[int, int] = self.align_patch_strands(dna, patch_group, strand_group,
                                                                      dna.scale_factor(p))
                erased_existing = dna.assign_patches_strands(strand_map)
                assert not erased_existing

            # check superimposition map
            m2 = np.stack([patch.position() for patch in p.patches()])
            m1 = np.stack([
                dna.strand_3p(dna.patch_strand_map[patch.get_id()]).pos
                for patch in p.patches()
            ])
            svd = SVDSuperimposer()
            svd.set(m1, m2)
            svd.run()
            relation.rms = svd.get_rms()
            self.check_rms(relation.rms, dna)

        print(f"Superimposed DNA origami on patchy particle type {p.type_id()} with RMS={relation.rms}")
        print(f"RMS / circumfrance: {relation.rms / (2 * math.pi * dna.center2patch_conf())}")
        assert p.num_patches() == len(dna.patch_strand_map)
        return PatchyOriRelation(p, dna, relation.rot, relation.perm, relation.rms)

    def align_patch_strands(self,
                            dna: DNAParticle,
                            patches: set[PLPatch],
                            strands: Iterable[int],
                            sf: float) -> dict[int, int]:
        """
        Aligns a group of patches (representing a single multidentate patch) with the 3' ends of strands
        using the Hungarian Algorithm for optimal pairing.
        Returns:
            a mapping where keys are ids of patches and values are strand ids
        """

        # get patch positions
        patch_positions = np.stack([patch.position() / sf for patch in patches])

        # get strand positions
        strand_positions = np.stack([dna.strand_3p(strand_id).pos for strand_id in strands])

        # Calculate the distance matrix between each patch and strand position
        distance_matrix = np.linalg.norm(patch_positions[:, None, :] - strand_positions[None, :, :], axis=2)

        # Use the Hungarian algorithm to find the optimal pairing
        patch_indices, strand_indices = linear_sum_assignment(distance_matrix)

        # Create the mapping of patch IDs to strand IDs based on the optimal pairing
        patch_ids = [patch.get_id() for patch in patches]
        strand_ids = list(strands)
        best_mapping = {patch_ids[i]: strand_ids[j] for i, j in zip(patch_indices, strand_indices)}

        return best_mapping

    def link_patchy_particle(self, p: PLPatchyParticle, dna: DNAParticle):
        """
        Links a patchy particle to a DNA structure, rotating the structure so that the orientation of the 3' ends of
        the patch strands match the position of the patches on the particle
        Note: PLEASE do not link multiple DNAParticle instances to the same patchy particle or vice
        versa!!!

        Parameters:
            p (PatchyBaseParticle): a patchy particle to link to a DNA particle
            dna (DNAParticle): a dna particle to link with the patchy particle
        """
        # assert len(self.patch_strand_ids) >= p.num_patches() or p.num_patches() % len(self.patch_strand_ids) == 0, \
        assert len(dna.flat_strand_ids) >= p.num_patches(), \
            f"Not enough patches on DNA particle ({len(dna.patch_strand_ids)}) to link DNA particle to " \
            f"on patchy particle({p.num_patches()})!"
        # compute strand mapping
        # if we have no strand map, construct one
        if not dna.has_strand_map():
            relation = self.match_patches_to_strands(p, dna)
        # if we have manually mapped strands to patches, we can just kinda ignore and move on
        else:
            relation = PatchyOriRelation(p, dna, np.identity(3), dna.patch_strand_map, 0)
        # apply the rotation to the dna structure so it matches the particle
        # dna.transform(relation.rot)
        # self.patch_strand_ids = [self.patch_strand_ids[i] for i in best_order[:-1]] # skip last position (centerpoint)
        dna.linked_particle = p

    def patch_group_coords(self, p: PLPatchyParticle, dna: DNAParticle) -> tuple[dict[int, list[np.ndarray]],
                                                                                 dict[int, list[np.ndarray]]]:
        # construct dict where keys are patch type ids and vals are lists of tooth patches
        pl_patch_coords: dict[int, list[np.ndarray]] = {}
        # do patch a1 vectors
        pl_patch_norms: dict[int, list[np.ndarray]] = {}
        # load each patch on the patchy particle
        for patch in p.patches():
            # if patch is multidentate, get src patch id as map key
            if self.patchy_scene.particle_types().is_multidentate():
                patch_type_id = self.patchy_scene.particle_types().get_src_map().get_src_patch(patch).type_id()
            else:
                # use patch id as map key, will end up with identity mappigng
                patch_type_id = patch.type_id()
            if patch_type_id not in pl_patch_coords:
                pl_patch_coords[patch_type_id] = list()
                pl_patch_norms[patch_type_id] = list()
            # scale patch with origami
            ppos = patch.position() / dna.scale_factor(p)
            pl_patch_coords[patch_type_id].append(ppos)
            # the pl a1 vector should always be a unit vector
            pnorm = patch.a1()
            pl_patch_norms[patch_type_id].append(pnorm)
        return pl_patch_coords, pl_patch_norms

    def align_patches(self, p: PLPatchyParticle, dna: DNAParticle) -> PatchyOriRelation:
        """
        Aligns patch centerpoints with the average positions of multidentate patches
        We can't do this strand-by-strand because if we do it will take until the sun goes out
        and then not work

        Parameters:

        Returns:
            a PatchyOriRelation object storing the patch order, rotation matrix, and rms
            to superimpose the DNA structure onto the patchy particle
        """
        sup = SVDSuperimposer()
        pl_patch_coords, pl_patch_norms = self.patch_group_coords(p, dna)

        # remember patch id order
        pid_order = list(pl_patch_coords.keys())
        assert len(pid_order) <= len(dna.patch_strand_ids), "Too many patches on patchy particle " \
                                                            "to map onto this DNA particle!"
        # compute centerpoints of each multidentate patch on the patchy particle
        pl_coords = np.stack([np.mean(plocs, axis=0)
                              for plocs in pl_patch_coords.values()])
        pl_norms = np.stack([np.mean(pnorms, axis=0) for pnorms in pl_patch_norms.values()])
        assert pl_coords.shape == (len(pid_order), 3), "Bad patch position matrix, somehow"
        # compute centers of mass of patches on DNA particle
        dna_patch_cmss = dna.get_patch_cmss()
        dna_patch_norms = dna.get_strand_group_norms()
        # lower rms = better so start with inf. rms
        best_rms = np.Inf
        best_order = None
        best_rot = None
        # compute center of mass of DNA particle
        cms = dna.cms()
        # loop permutations of patche strand gorups
        m2 = np.concatenate([pl_coords,
                             pl_norms,
                             np.zeros(shape=(1, 3))], axis=0)
        for perm in itertools.permutations(range(len(dna.patch_strand_ids)), r=len(pid_order)):
            # configure svd superimposer
            # m1 = generate matrix of origami patches, norms, centerpoint
            m1 = np.concatenate([dna_patch_cmss[perm, :],
                                 dna_patch_norms[perm, :],
                                 cms[np.newaxis, :]], axis=0)
            # m2 = patchy particle patches + centerpoint

            if m1.shape != m2.shape:
                raise Exception(f"Mismatch between patch position matrix on pl particle ({m1.shape}) "
                                f"and on dna particle ({m2.shape})")
            sup.set(m1, m2)

            # FIRE
            sup.run()
            # if rms is better
            if sup.get_rms() < best_rms:
                assert abs(np.linalg.det(sup.rot) - 1) < 1e-8
                # save order
                best_order = dict(zip(pid_order, perm))
                best_rot = sup.rot
                # save rms
                best_rms = sup.get_rms()

        assert best_order is not None

        diff = pl_coords @ best_rot - dna_patch_cmss[np.array(list(best_order.values()))]
        rmsd = np.linalg.norm(diff)

        # check patch normals
        # for pl_patch_id, strand_id in best_order.items():

        return PatchyOriRelation(p, dna, best_rot, best_order, best_rms)

    def ready_to_position(self) -> bool:
        return all([ptype.get_type() in self.particle_type_map
                    for ptype in self.patchy_scene.particle_types()])

    def position_particles(self):
        """
        Positions particles?
        IDK
        """
        assert self.ready_to_position()
        self.dna_particles = dict()
        # get scene particles
        particles: list[PLPatchyParticle] = self.patchy_scene.particles()
        pl = len(particles)
        for i, particle in enumerate(particles):
            # clone dna particle
            origami: DNAParticle = deepcopy(self.get_dna_origami(particle))
            # origami: DNAParticle = self.get_dna_origami(particle).clone()
            print(f"{i}/{pl}", end="\r")
            assert origami.linked_particle.matches(particle)
            # align dna particle with patchy
            origami.instance_align(particle)
            scale_factor = self.get_scale_factor()
            # scale factor tells us how to convert MGL distance units into our DNA model distances
            # transform origami
            origami.transform(tran=particle.position() * scale_factor)

            # we finished the positioning
            self.dna_particles[particle.get_uid()] = origami
        print()

    def bind_particles3p(self):
        """
        Creates double-stranded linkers to bind particles together
        """

        # use pl function to iter bound particles
        for p1, p2 in self.patchy_scene.iter_bound_particles():
            p1_dna = self.get_dna_particle(p1)
            p2_dna = self.get_dna_particle(p2)
            # iter pairs of patches that connect p1 to p2
            for patch1, patch2 in self.patchy_scene.iter_binding_patches(p1, p2):
                self.bind_patches_3p(p1_dna,
                                     patch1,
                                     p2_dna,
                                     patch2)

    def bind_patches_3p(self,
                        particle1: DNAParticle,
                        patch1: PLPatch,
                        particle2: DNAParticle,
                        patch2: PLPatch):
        """
        Binds two patches together by adding sticky ends at the 3' ends.
        Parameters:
            particle1 (DNAParticle): DNAParticle object representing particle a
            patch1 (MGLPatch): mgl patch object?
            particle2 (DNAParticle): DNAParticle object representing particle b
            patch2 (MGLPatch): mgl patch object?
        """
        # chcek that patchs are free
        assert ((not self.is_track_conf_stickies()) or particle1.linked_particle.get_uid(), patch1.type_id()) not in self.sticky_book
        assert ((not self.is_track_conf_stickies()) or particle2.linked_particle.get_uid(), patch2.type_id()) not in self.sticky_book

        # find nucleotides which will be extended to form patches
        assert patch1.color() + patch2.color() == 0

        start_position1 = particle1.patch_3p(patch1).pos
        start_position2 = particle2.patch_3p(patch2).pos

        start_vector_1 = normalize(start_position2 - start_position1)

        patch_distance = dist(start_position1, start_position2)
        # length of dna in oxdna units
        dna_len = (2 * self.spacer_length + len(self.color_sequence(patch1.color()))) * BASE_BASE

        # check distances
        if patch_distance > 2 * self.get_expected_pad_distance():
            print(f"Distance between patches = {patch_distance}. "
                  f"Expected distance {self.get_expected_pad_distance()}")

        assert not self.check_strict_rc or self.color_sequence(patch1.color()) == rc(
            self.color_sequence(patch2.color()))

        # retrieve sequences from map + add spacers
        patch1_seq = self.spacer_length * "T" + self.color_sequence(patch1.color())
        patch2_seq = self.spacer_length * "T" + self.color_sequence(patch2.color())

        strand1, strand2 = construct_strands(patch1_seq + self.spacer_length * "T",
                                             # need to add fake spacer to make code no go boom
                                             start_position1 + start_vector_1 * ((patch_distance - dna_len) / 2),
                                             start_vector_1,
                                             rbases=patch2_seq)
        strand1 = strand1[:len(patch1_seq)]  # shave off dummy spacer from previous line of code

        # extend strands
        particle1.patch_strand(patch1).prepend(strand1[::-1])
        particle2.patch_strand(patch2).prepend(strand2[::-1])

        # register stickies in sticky book
        if self.is_track_conf_stickies():
            self.sticky_book.add((particle1.linked_particle.get_uid(), patch1.type_id()))
            self.sticky_book.add((particle2.linked_particle.get_uid(), patch2.type_id()))

        self.bondcount += 1

    def add_unbound_stickies(self):
        """
        there is  a better way to handle this stuff but in the current stage of development,
        code stability is a top priorituy
        """
        # iter individual particles
        assert self.is_track_conf_stickies()
        for particle in self.get_particles():
            # iter patches
            for patch in particle.linked_particle.patches():
                # check that we have not already 3'-extended this SE
                if not (particle.linked_particle.get_uid(), patch.type_id()) in self.sticky_book:
                    self.add_patch_sticky(particle, patch)

    def add_patch_sticky(self, particle: DNAParticle, patch: PLPatch):
        """
        extends a staple to add a single-stranded overhang (aka sticky end) without using it to bind to
        another particle
        TODO: merge with bind_patches_3p
        """
        assert ((not self.is_track_conf_stickies()) or particle.linked_particle.get_uid(), patch.type_id()) not in self.sticky_book

        # find strand 3' base start position
        start_position = particle.patch_3p(patch).pos

        # use patch info to find direction vector
        direction_vector = particle.linked_particle.patch_a1(patch)

        # get sequence & add spacer
        patch1_seq = self.spacer_length * "T" + self.color_sequence(patch.color())

        # construct DNA strand
        strand, _ = construct_strands(patch1_seq,
                                      # start at 1 base past the start point
                                      start_position + direction_vector * BASE_BASE,
                                      direction_vector
                                      )
        particle.patch_strand(patch).prepend(strand[::-1])
        if self.is_track_conf_stickies():
            self.sticky_book.add((particle.linked_particle.get_uid(), patch.type_id()))

    def get_particles(self) -> list[DNAParticle]:
        """
        Returns DNA structures that correspond to particles
        """
        if self.dna_particles is None:
            print("positioning particles")
            self.position_particles()
        return list(self.dna_particles.values())

    def get_dna_particle(self, pl: Union[int, PLPatchyParticle]) -> DNAParticle:
        """
        Gets the DNA particle instace (not type) for a PL particle instance
        """
        # position particles if missing
        self.get_particles()
        if isinstance(pl, PLPatchyParticle):
            return self.get_dna_particle(pl.get_uid())
        else:
            assert isinstance(pl, int)
            assert pl in self.dna_particles
            return self.dna_particles[pl]

    def convert(self, unbound_stickies: bool = True):
        """
        Converts a scene containing joined MGL particles to an oxDNA model consisting of
        DNA origamis joined by sticky end handles.
        """

        print("binding particles using 3p patches")
        self.bind_particles3p()
        assert self.expected_num_edges == -1 or self.bondcount == self.expected_num_edges, \
            f"Wrong number of bonds created! Expected {self.expected_num_edges} bonds, got {self.bondcount}."
        if self.is_track_conf_stickies() and unbound_stickies:
            self.add_unbound_stickies()
        print("Done!")

    def dump_monomers(self, fp: Union[None, Path, str] = None, only_present=False):
        """
        Saves all dna particle types in their own .top and .dat files
        """
        # handle inputs
        if fp is None:
            # default to drop files in output dir
            fp = get_output_dir()
        elif isinstance(fp, str):
            fp = Path(fp)
            if not fp.is_absolute():
                fp = get_output_dir() / fp
        if not fp.exists():
            os.makedirs(fp)
        for ptype in self.particle_type_map.values():
            if not only_present or self.patchy_scene.particle_type_counts()[ptype.linked_particle.get_type()] > 0:
                self.dump_monomer(ptype, fp)

    def dump_monomer(self, ptype: DNAParticle, fp: Path):
        assert ptype.has_linked(), "Cannot dump monomer for unlinked DNA particle"
        if np.abs(ptype.linked_particle.rotation() - np.identity(3)).sum() > 1e-6:
            print("Warning: Particle has non-identity rotation! "
                  "Patch a1s are going to be messed up! But you can technically proceed")
        cpy = copy.deepcopy(ptype)
        # clone to add strands
        # iter strand id map
        for patch_id in cpy.patch_strand_map:
            # particle types don't have sticky ends added until they're positioned so we need to add them
            # here before we export
            patch = cpy.linked_particle.patch_by_id(patch_id)
            seq = "T" * self.spacer_length + self.color_sequence(patch.color())
            # last-base a1 is a really bad method for doing strand vector, compute patch a1 instead
            strand1, _ = construct_strands(seq,
                                           start_pos=cpy.patch_strand(patch_id)[0].pos,
                                           helix_direction=patch.a1())

            cpy.patch_strand(patch_id).prepend(strand1[::-1])

        cpy.export_top_conf(fp / (cpy.linked_particle.name() + ".top"),
                            fp / (cpy.linked_particle.name() + ".dat"))

    def save_top_dat(self, write_top_path: Union[Path, str], write_conf_path: Union[Path, str]):
        if isinstance(write_top_path, str):
            write_top_path = Path(write_top_path)
        if not write_top_path.is_absolute():
            write_top_path = get_output_dir() / write_top_path

        if isinstance(write_conf_path, str):
            write_conf_path = Path(write_conf_path)
        if not write_conf_path.is_absolute():
            write_conf_path = get_output_dir() / write_conf_path

        assert write_top_path.parent.exists()
        assert write_conf_path.parent.exists()
        if write_top_path.parent != write_conf_path.parent:
            print("You're technically allowed to do this but I do wonder why")

        merged_conf = self.get_full_conf()

        merged_conf.export_top_conf(write_top_path, write_conf_path)
        print(f"Wrote topopogy file to {str(write_top_path)}")

        print(f"Wrote conf file to {str(write_conf_path)}")

    def save_oxview(self, write_oxview_path: Union[Path, str], clusters=False, color_by_type=False):
        """
        Saves the converted structure as an .oxview file
        Parameters:
            write_oxview_path the path at which to write the oxview file
            clusters whether to save each DNA particle in the oxview file as its own cluster (will be slower)
            color_by_type whether to color the DNA particles according to patchy particle type (will be slower)
        """
        if isinstance(write_oxview_path, str):
            write_oxview_path = Path(write_oxview_path)
        if not write_oxview_path.is_absolute():
            write_oxview_path = get_output_dir() / write_oxview_path

        merged_conf = self.get_full_conf(clusters=clusters or color_by_type)
        oxview_json = merged_conf.get_oxview_json()
        # colorcode particles by type
        if color_by_type:
            for strand in oxview_json['systems'][0]['strands']:
                for monomer in strand['monomers']:
                    particle = self.dna_particles[monomer['cluster']]
                    particle_type = particle.linked_particle.type_id()
                    color = selectColor(particle_type)  # get hex string code of color
                    # oxview file format requires we convert the hex code to base 10
                    monomer["color"] = int(color[1:], 16)
        with write_oxview_path.open("w") as f:
            json.dump(oxview_json, f, indent=4)
        print(f"Wrote OxView file to {str(write_oxview_path)}")

    def iter_sticky_staples(self,
                            pl_type: Union[int, PLPatchyParticle, None] = None,
                            incl_no_sticky: bool = True,
                            incl_absent: bool = True) -> Generator[tuple[DNAParticle, int,
                                                                         int, DNAStructureStrand],
                                                                   None,
                                                                   None]:
        if pl_type is None:
            for p in self.patchy_scene.particle_types():
                if incl_absent or self.patchy_scene.particle_type_counts()[p.get_type()] > 0:
                    yield from self.iter_sticky_staples(p.get_type())
        elif isinstance(pl_type, PLPatchyParticle):
            yield from self.iter_sticky_staples(pl_type.get_type())
        else:
            print(f"Particle {pl_type}")
            ptype = self.particle_type_map[pl_type]
            assert ptype.has_linked(), "Cannot save stickies for unliked particle"
            if np.abs(ptype.linked_particle.rotation() - np.identity(3)).sum() > 1e-6:
                print("Warning: Particle has non-identity rotation! "
                      "Patch a1s are going to be messed up! But you can technically proceed")
            # clone to add strands
            # iter strand id map
            for strand_id in ptype.flat_strand_ids:
                strand = copy.deepcopy(ptype.strands[strand_id])
                # particle types don't have sticky ends added until they're positioned so we need to add them
                # here before we export
                patch_id = ptype.patch_for_strand(strand_id)

                if patch_id is not None:
                    patch = ptype.linked_particle.patch_by_id(patch_id)
                    seq = "T" * self.spacer_length + self.color_sequence(patch.color())
                    # last-base a1 is a really bad method for doing strand vector, compute patch a1 instead
                    strand1, _ = construct_strands(seq,
                                                   start_pos=ptype.patch_strand(patch_id)[0].pos,
                                                   helix_direction=patch.a1())
                    strand.prepend(strand1[::-1])
                elif not incl_no_sticky:
                    continue
                yield ptype, strand_id, patch_id, strand

    def print_sticky_staples(self,
                             pl_type: Union[int, PLPatchyParticle, None] = None,
                             incl_no_sticky: bool = True,
                             incl_absent: bool = True):
        for dna, strand_id, patch_id, strand in self.iter_sticky_staples(pl_type, incl_no_sticky, incl_absent):
            sz = f"Strand {strand_id} : 5' {strand.seq(True)} 3'"
            if patch_id is not None:
                print(f"Patch {patch_id} : " + sz)
            else:
                print(sz)

    def export_stickys_staples(self,
                               fp: Union[Path, str],
                               by_row=True,
                               incl_no_sticky: bool = True,
                               incl_absent: bool = True,
                               incl_original_patch_nums: bool = False):
        """
        Exports sticky-end staples to an excel file for ordering
        Parameters:
            by_row: if true, the excel spreadsheet will export so as to make each row on the 96-well plate (each pair of letters) correspond to a specific particle. this will go. very poorly if each DNA particle doesn't have exactly 24 sticky candidates
            incl_no_sticky: if true, the spreadsheet will only include strands which have sticky-ends (the strands that facilitate interparticle interaction)
            incl_absent: if true, the converted will include all staples required for the origiami, not just
            incl_original_patch_nums: if true, the converter will include the patch numbers or (for palycuebes) direction names. This currently only works with polycubes.
        """
        if isinstance(fp, str):
            fp = Path(fp)
        assert fp.parent.exists(), f"No such location as {str(fp.parent)}"
        wb = openpyxl.Workbook()
        ws = wb.active

        # TODO: CHECK USING PATCHY CONVERT PIPELINE OBJECT TYPES
        is_polycube = True

        # verify validity of by_row parameter
        assert not by_row or all([len(particle.flat_strand_ids) == 24 for particle in self.particle_type_map.values()]), \
            "Option `by_row` is not supported for particles with a number of binding strands other than 24."

        idx = 1
        ws[f"A{idx}"] = "Wells"
        ws[f"B{idx}"] = "Name"
        ws[f"C{idx}"] = "Sequences"
        idx += 1
        prev_pid = None
        for i, (dna, strand_id, patch_id, strand) in enumerate(self.iter_sticky_staples(incl_no_sticky=incl_no_sticky,
                                                                                        incl_absent=incl_absent)):
            if patch_id is None and not incl_no_sticky:
                continue
            if prev_pid is None:
                prev_pid = dna.linked_particle.get_type()
            if by_row and dna.linked_particle.get_type() != prev_pid:
                # if each particle type should be its own row on the sheet
                if (idx - 2) % 12 != 0:
                    idx = 12 * math.ceil((idx - 2) / 12) + 2
                prev_pid = dna.linked_particle.get_type()

            # trust me
            if idx > 97:
                ws = wb.create_sheet()
                idx = 1

                ws[f"A{idx}"] = "Wells"
                ws[f"B{idx}"] = "Name"
                ws[f"C{idx}"] = "Sequences"
                idx += 1
            plate_cell = idx - 2

            plate_row = "ABCDEFGH"[int(plate_cell / 12)]
            plate_col = plate_cell % 12 + 1

            ws[f"A{idx}"] = f"{plate_row}{plate_col}"
            seq_string_descriptor = f"ParticleType{dna.linked_particle.get_type()}"
            seq_string_descriptor += f"_Strand{strand_id}"

            if patch_id is not None:
                seq_string_descriptor += f"_Patch{patch_id}"
                if incl_original_patch_nums:
                    patch_src = self.patchy_scene.particle_types().get_src_map().get_src_patch(
                        self.patchy_scene.particle_types().patch(patch_id))
                    if is_polycube:
                        # disgusting hack which i hate
                        # todo: burn it with fire
                        patch_direction_index = diridx(patch_src.position() / np.linalg.norm(patch_src.position()))
                        assert -1 < patch_direction_index < len(FACE_NAMES)
                        seq_string_descriptor += f"_{FACE_NAMES[patch_direction_index]}"
            ws[f"B{idx}"] = seq_string_descriptor
            assert strand.seq(True)  # unclear when this would ever fire, doubel check
            ws[f"C{idx}"] = strand.seq(True)
            idx += 1

        wb.save(fp)
